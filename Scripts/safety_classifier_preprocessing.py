"""
Safety Classifier Preprocessing Script
=======================================
Reads the Summary sheet from main.xlsx which contains the
Calculated Hazard Severity (CHS) score per cell, as defined by:

    Lin et al., "Mechanically induced thermal runaway severity analysis
    for Li-ion batteries", Journal of Energy Storage 61 (2023).

The CHS score (0-100) is computed by the original dataset authors using:
    - Cell temperature
    - Rate of temperature increase
    - Cell capacity
    - State of charge (SOC)
    - Voltage drop
    - Voltage drop rate
Each term is weighted and scaled to 0-100.

This script uses CHS directly as the label — no heuristic voltage
thresholding. For the RL reward function P(runaway), CHS/100 gives
a continuous probability. A binary label (CHS >= 50) is also included.

Reference:
    Lin L. et al. (2023). DOI: 10.1016/j.est.2022.106798
    Dataset: https://data.mendeley.com/datasets/sn2kv34r4h/2

Output:
    safety_classifier_dataset.csv
"""

import os
import re
import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────
# !! SET YOUR PATHS HERE !!
# ──────────────────────────────────────────────────────────────────

# Path to main.xlsx (the summary file from the dataset)
MAIN_XLSX_PATH = r"C:\Users\saumy\OneDrive\Desktop\IN Pallavi ma'am\Datasets extracted\Safety classifer\Mechanically Induced Thermal Runaway for Li-ion Batteries\Mechanically Induced Thermal Runaway for Li-ion Batteries\main.xlsx"

# Output CSV path — saved in same folder as main.xlsx
OUTPUT_CSV = r"safety_classifier_dataset.csv"

# ──────────────────────────────────────────────────────────────────
# CHEMISTRY MAP  (used as contextual variable θ in MDP)
# ──────────────────────────────────────────────────────────────────
CHEMISTRY_MAP = {'LCO': 0, 'LFP': 1, 'NMC': 2, 'NCA': 3}


def parse_chemistry(filename):
    fn = str(filename).upper()
    for chem in ['LCO', 'LFP', 'NMC', 'NCA']:
        if chem in fn:
            return chem
    return 'UNKNOWN'


def parse_soc(filename):
    match = re.search(r'(\d+)SOC', str(filename), re.IGNORECASE)
    return int(match.group(1)) if match else None


def parse_capacity(filename):
    fn = str(filename)
    match = re.search(r'(\d+(?:\.\d+)?)\s*Ah', fn, re.IGNORECASE)
    if match:
        val = float(match.group(1))
        full = match.group(0).lower()
        return val if 'mah' in full else val * 1000
    return None


def read_summary(main_xlsx_path):
    """
    Summary columns (0-indexed):
        0  = FileName
        1  = OHS text description
        2  = OHS numeric (1-7)
        3  = CHS score (0-100, continuous)  ← PRIMARY LABEL
        4  = Capacity (mAh)
        5  = SOC (%)
        9  = max T (°C)
        10 = rate T (°C/s)
        11 = voltage difference (V)
    """
    wb = load_workbook(main_xlsx_path, read_only=True)
    ws = wb['Summary']
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[1:]:
        filename  = row[0]
        ohs_text  = row[1]
        ohs_score = row[2]
        chs_score = row[3]
        capacity  = row[4]
        soc       = row[5]
        t_max     = row[9]
        t_rate    = row[10]
        v_diff    = row[11]

        if filename is None or chs_score is None:
            continue
        if not isinstance(chs_score, (int, float)):
            continue

        chemistry = parse_chemistry(filename)
        if capacity is None:
            capacity = parse_capacity(filename)
        if soc is None:
            soc = parse_soc(filename)

        try:
            ohs_numeric = int(float(ohs_score)) if ohs_score is not None else None
        except (ValueError, TypeError):
            ohs_numeric = None

        records.append({
            'filename':       filename,
            'chemistry':      chemistry,
            'chemistry_id':   CHEMISTRY_MAP.get(chemistry, -1),
            'capacity_mAh':   capacity,
            'soc':            soc,
            'chs_score':      round(float(chs_score), 4),
            'p_runaway':      round(float(chs_score) / 100, 4),
            'runaway_binary': int(float(chs_score) >= 50),
            'ohs_numeric':    ohs_numeric,
            'ohs_text':       ohs_text,
            't_max_C':        round(float(t_max),  4) if isinstance(t_max,  (int, float)) else None,
            't_rate_Cs':      round(float(t_rate), 6) if isinstance(t_rate, (int, float)) else None,
            'v_diff_V':       round(float(v_diff), 4) if isinstance(v_diff, (int, float)) else None,
        })

    return pd.DataFrame(records)


def main():
    print(f"Reading: {MAIN_XLSX_PATH}")
    df = read_summary(MAIN_XLSX_PATH)

    # Keep only the three target chemistries
    before = len(df)
    df = df[df['chemistry'].isin(['LCO', 'LFP', 'NMC'])].reset_index(drop=True)
    print(f"Filtered out {before - len(df)} non-target cells (kept LCO, LFP, NMC only)")

    print(f"\n{'='*55}")
    print(f"  Total cells processed : {len(df)}")
    print(f"  CHS range             : {df['chs_score'].min():.1f} - {df['chs_score'].max():.1f}")
    print(f"  Binary runaway (>=50) : {df['runaway_binary'].sum()} / {len(df)} "
          f"({df['runaway_binary'].mean()*100:.1f}%)")
    print(f"{'='*55}")

    print("\nChemistry breakdown:")
    print(df.groupby('chemistry')[['chs_score','runaway_binary']].agg(
        count=('chs_score','count'),
        chs_mean=('chs_score','mean'),
        runaway_count=('runaway_binary','sum')
    ).round(1).to_string())

    print("\nSOC breakdown:")
    print(df.groupby('soc')[['chs_score','runaway_binary']].agg(
        count=('chs_score','count'),
        chs_mean=('chs_score','mean'),
        runaway_count=('runaway_binary','sum')
    ).round(1).to_string())

    df.to_csv(OUTPUT_CSV, index=False)
    print(f"\n  Saved to: {OUTPUT_CSV}")
    print("\nColumn guide:")
    print("  chs_score      -> continuous label 0-100  (regression target)")
    print("  p_runaway      -> chs_score/100            (P(runaway) in RL reward)")
    print("  runaway_binary -> 1 if chs_score >= 50     (binary classifier target)")
    print("  chemistry_id   -> 0=LCO, 1=LFP, 2=NMC     (contextual variable theta)")


if __name__ == '__main__':
    main()